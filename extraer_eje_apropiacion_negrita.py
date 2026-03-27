import openpyxl
import pandas as pd

def extraer_eje_apropiacion_negrita(xlsx_path, sheet_name=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Buscar encabezado de la columna APROPIACION VIGENTE DEP.GSTO.
    ap_header_normalized = "APROPIACION VIGENTE DEP.GSTO."
    header_row = None
    ap_col_idx = None
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=30):
        row_headers = []
        for idx, cell in enumerate(row):
            val = str(cell.value).replace('\n', ' ').replace('\r', ' ').replace('  ', ' ').strip().upper() if cell.value else ''
            row_headers.append(val)
            if ap_header_normalized in val.replace('.', '').replace('  ', ' '):
                header_row = cell.row
                ap_col_idx = idx
        if header_row is not None:
            headers = row_headers
            break
    if header_row is None or ap_col_idx is None:
        raise Exception("No se encontró la columna 'APROPIACION VIGENTE DEP.GSTO.' en el archivo.")

    # Extraer datos en negrita de la columna AP
    datos = []
    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
        cell_ap = row[ap_col_idx]
        if cell_ap.value is not None and hasattr(cell_ap, 'font') and cell_ap.font and cell_ap.font.bold:
            fila = {headers[ap_col_idx]: cell_ap.value}
            datos.append(fila)
    df = pd.DataFrame(datos)
    return df

if __name__ == "__main__":
    import sys
    archivo = sys.argv[1] if len(sys.argv) > 1 else "930810-EJE-27-11-2025.xlsx"
    df = extraer_eje_apropiacion_negrita(archivo)
    print("Valores en negrita de la columna APROPIACION VIGENTE DEP.GSTO. (AP):")
    print(df)
    df.to_csv("apropiacion_vigente_negrita_extraida.csv", index=False)
    print("\nArchivo 'apropiacion_vigente_negrita_extraida.csv' generado.")
