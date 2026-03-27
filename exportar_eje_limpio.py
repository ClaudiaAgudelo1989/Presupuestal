import openpyxl
import csv

# Configuración
xlsx_path = "930810-EJE-27-11-2025.xlsx"
csv_salida = "eje_limpio.csv"

# Nombres de columnas según la tabla MySQL
columnas = [
    "tipo", "cta", "subc", "objg", "ord", "sord", "item", "sitem", "concepto", "fuente", "situacion", "rec", "recurso",
    "apropiacion_vigente_dep_gsto", "total_cdp_dep_gstos", "apropiacion_disponible_dep_gsto", "total_cdp_modificacion_dep_gstos",
    "total_compromiso_dep_gstos", "cdp_por_comprometer_dep_gstos", "total_obligaciones_dep_gstos", "compromiso_por_obligar_dep_gstos",
    "total_ordenes_pago_dep_gstos", "obligaciones_por_ordenar_dep_gstos", "pagos_dep_gstos", "ordenes_pago_por_pagar_dep_gstos",
    "total_reintegros_dep_gstos"
]

# Carga el archivo Excel
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb.active

# Busca la fila de encabezados
header_map = {}
for row in ws.iter_rows(min_row=1, max_row=5):
    for idx, cell in enumerate(row):
        if cell.value:
            header_map[str(cell.value).strip().replace('\n', ' ')] = idx
    if len(header_map) >= len(columnas):
        break

# Mapeo manual de encabezados de Excel a columnas de la tabla (ajusta si es necesario)
excel_to_db = {
    "TIPO": "tipo",
    "CTA": "cta",
    "SUBC": "subc",
    "OBJG": "objg",
    "ORD": "ord",
    "SORD": "sord",
    "ITEM": "item",
    "SITEM": "sitem",
    "CONCEPTO": "concepto",
    "FUENTE": "fuente",
    "SITUACION": "situacion",
    "REC.": "rec",
    "RECURSO": "recurso",
    "APROPIACION VIGENTE DEP.GSTO.": "apropiacion_vigente_dep_gsto",
    "TOTAL CDP DEP.GSTOS": "total_cdp_dep_gstos",
    "APROPIACION N DISPONIBLE DEP.GSTO.": "apropiacion_disponible_dep_gsto",
    "TOTAL CDP MODIFICACION DEP.GSTOS": "total_cdp_modificacion_dep_gstos",
    "TOTAL COMPROMISO DEP.GSTOS": "total_compromiso_dep_gstos",
    "CDP POR COMPROMETER DEP.GSTOS": "cdp_por_comprometer_dep_gstos",
    "TOTAL OBLIGACIONES DEP.GSTOS": "total_obligaciones_dep_gstos",
    "COMPROMISO POR OBLIGAR DEP.GSTOS": "compromiso_por_obligar_dep_gstos",
    "TOTAL ORDENES DE PAGO DEP.GSTOS": "total_ordenes_pago_dep_gstos",
    "OBLIGACIONES POR ORDENAR DEP.GSTOS": "obligaciones_por_ordenar_dep_gstos",
    "PAGOS DEP.GSTOS": "pagos_dep_gstos",
    "ORDENES DE PAGO POR PAGAR DEP.GSTOS": "ordenes_pago_por_pagar_dep_gstos",
    "TOTAL REINTEGROS DEP.GSTOS": "total_reintegros_dep_gstos"
}

# Encuentra la fila de encabezados real
def find_header_row():
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10)):
        values = [str(cell.value).strip().replace('\n', ' ') if cell.value else "" for cell in row]
        if all(excel_to_db.get(v) for v in values if v):
            return i + 1
    return None

header_row_idx = find_header_row()
if header_row_idx is None:
    print("No se encontró la fila de encabezados correcta.")
    exit(1)

# Exporta los datos
with open(csv_salida, mode="w", newline='', encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(columnas)
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
        fila = []
        for excel_col, db_col in excel_to_db.items():
            idx = header_map.get(excel_col)
            cell = row[idx] if idx is not None and idx < len(row) else None
            if db_col.startswith("apropiacion") or db_col.startswith("total") or db_col.startswith("cdp") or db_col.startswith("pagos") or db_col.startswith("ordenes") or db_col.startswith("compromiso") or db_col.startswith("obligaciones"):
                # Números: si vacío, pon 0
                valor = cell.value if cell and cell.value not in (None, "") else 0
                try:
                    valor = float(str(valor).replace(",", "").replace(" ", "")) if valor != 0 else 0
                except Exception:
                    valor = 0
                fila.append(valor)
            else:
                # Texto: si vacío, pon ""
                valor = cell.value if cell and cell.value not in (None, "") else ""
                fila.append(str(valor).strip())
        writer.writerow(fila)

print(f"Archivo CSV exportado: {csv_salida}")
