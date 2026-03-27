import openpyxl
import pandas as pd

def build_eje_table_bold_only(xlsx_path, sheet_name=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Buscar la fila de encabezados y la columna de apropiación vigente (soporta saltos de línea y espacios)
    header_row = None
    col_aprop = None
    ap_header_normalized = "APROPIACION VIGENTE DEP.GSTO."
    for row in ws.iter_rows(min_row=1, max_row=30):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                normalized = str(cell.value).replace('\n', ' ').replace('\r', ' ').replace('  ', ' ').strip().upper()
                if ap_header_normalized in normalized.replace('  ', ' ').replace('.', '').replace('  ', ' '):
                    header_row = cell.row
        if header_row:
            break
    if not header_row:
        raise Exception("No se encontró la fila de encabezados para 'APROPIACION VIGENTE DEP.GSTO.'")

    # Leer encabezados normalizando saltos de línea y espacios
    headers = []
    ap_col_idx = None
    for idx, cell in enumerate(ws[header_row]):
        if cell.value:
            normalized = str(cell.value).replace('\n', ' ').replace('\r', ' ').replace('  ', ' ').strip().upper()
            headers.append(normalized)
            if ap_header_normalized in normalized.replace('  ', ' ').replace('.', '').replace('  ', ' '):
                ap_col_idx = idx
        else:
            headers.append(None)
    if ap_col_idx is None:
        raise Exception("No se encontró la columna 'APROPIACION VIGENTE DEP.GSTO.' en los encabezados")

    data = []
    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
        cell_aprop = row[ap_col_idx]
        if cell_aprop.value is not None and hasattr(cell_aprop, 'font') and cell_aprop.font and cell_aprop.font.bold:
            values = [cell.value for cell in row]
            data.append(values)
    df = pd.DataFrame(data, columns=headers)
    return df
