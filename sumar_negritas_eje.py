from openpyxl import load_workbook
import csv


# CONFIGURACIÓN
xlsx_path = "930810-EJE-27-11-2025.xlsx"  # Cambia aquí el nombre del archivo Excel
# Usa el encabezado exacto proporcionado por el usuario, con salto de línea real
nombre_columna_aprop = "APROPIACION\nVIGENTE DEP.GSTO."
nombre_columna_centro = "RECURSO"
csv_salida = "apropiacion_vigente_negrita.csv"

wb = load_workbook(xlsx_path, data_only=True)
ws = wb.active

# Buscar la fila de encabezados y las columnas requeridas


# Buscar la fila de encabezados y las columnas requeridas
header_row = None
col_aprop = None
col_centro = None
for row in ws.iter_rows(min_row=1, max_row=30):
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            # Normaliza saltos de línea y compara exactamente
            valor_celda = cell.value.replace("\r\n", "\n").replace("\r", "\n")
            if valor_celda == nombre_columna_aprop:
                header_row = cell.row
                col_aprop = cell.column
            if cell.value.strip().upper() == nombre_columna_centro.upper():
                col_centro = cell.column
    if header_row and col_aprop and col_centro:
        break
if not header_row or not col_aprop:
    print(f"No se encontró la columna '{nombre_columna_aprop}'")
    exit(1)
if not col_centro:
    print(f"No se encontró la columna '{nombre_columna_centro}'")
    exit(1)



# Recorrer filas y recolectar solo celdas en negrita y numéricas
datos_exportar = []
suma_total = 0
por_centro = {}
for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
    cell_aprop = row[col_aprop-1]
    cell_centro = row[col_centro-1]
    if cell_aprop.value is not None and hasattr(cell_aprop, 'font') and cell_aprop.font and cell_aprop.font.bold:
        try:
            valor = float(cell_aprop.value)
        except Exception:
            continue
        centro = cell_centro.value if cell_centro.value else "(Sin centro)"
        datos_exportar.append([centro, valor])
        suma_total += valor
        por_centro[centro] = por_centro.get(centro, 0) + valor

# Exportar a CSV
with open(csv_salida, mode="w", newline='', encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow([nombre_columna_centro, nombre_columna_aprop])
    writer.writerows(datos_exportar)

print(f"TOTAL APROPIACION VIGENTE (solo negritas): {suma_total:,.2f}")
print(f"Archivo CSV exportado: {csv_salida}")
print("\nDesglose por centro:")
for centro, valor in por_centro.items():
    print(f"- {centro}: {valor:,.2f}")
