import openpyxl
import pandas as pd

def inspect_excel_structure(xlsx_path, sheet_name=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    # Buscar encabezados y mostrar nombres de columnas
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=30):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "APROPIACION" in cell.value.upper():
                header_row = cell.row
                break
        if header_row:
            break
    if not header_row:
        print("No se encontró la fila de encabezados.")
        return
    headers = [cell.value for cell in ws[header_row]]
    print(f"Fila de encabezados: {header_row}")
    print(f"Encabezados detectados: {headers}")
    # Mostrar las primeras filas de la columna AP
    ap_col_idx = None
    for idx, h in enumerate(headers):
        if h and "APROPIACION" in str(h).upper():
            ap_col_idx = idx
            break
    if ap_col_idx is not None:
        print(f"Columna APROPIACION VIGENTE DEP.GSTO. (AP): índice {ap_col_idx}")
        for i, row in enumerate(ws.iter_rows(min_row=header_row+1, max_row=header_row+11)):
            val = row[ap_col_idx].value
            bold = getattr(row[ap_col_idx].font, 'bold', False)
            print(f"Fila {header_row+1+i}: valor={val}, negrita={bold}")
    else:
        print("No se encontró la columna APROPIACION VIGENTE DEP.GSTO.")
